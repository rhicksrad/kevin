import json
import re
import zipfile
from datetime import UTC, datetime, time
from pathlib import Path
from typing import Any, Dict, List, Optional
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.utils.cell import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
GRID_PATH = ROOT / "2025 Grid.xlsm"
OUTPUT_PATH = ROOT / "docs" / "assets" / "grid-data.json"


def serialise_datetime(value: Any) -> Optional[str]:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if value is None:
        return None
    return str(value)


SCHEDULE_START_COLUMN = 8
SCHEDULE_GROUP_WIDTH = 3
SCHEDULE_HEADER_TOKENS = {
    "rk",
    "team",
    "teams",
    "time",
    "circa",
}


MAIN_NAMESPACE = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


class WorkbookStyleInspector:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self.archive: Optional[zipfile.ZipFile] = None
        self.sheet_targets: Dict[str, str] = {}
        self.style_outcome: Dict[int, str] = {}
        self._style_cache: Dict[str, Dict[str, int]] = {}

    def __enter__(self) -> "WorkbookStyleInspector":
        self.archive = zipfile.ZipFile(self.workbook_path)
        self._load_metadata()
        return self

    def __exit__(self, exc_type, exc, traceback) -> None:  # type: ignore[override]
        if self.archive is not None:
            self.archive.close()
            self.archive = None

    def _load_metadata(self) -> None:
        assert self.archive is not None

        workbook_tree = ET.fromstring(self.archive.read("xl/workbook.xml"))
        rels_tree = ET.fromstring(self.archive.read("xl/_rels/workbook.xml.rels"))

        relationship_map = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_tree
            if "Id" in rel.attrib and "Target" in rel.attrib
        }

        sheets_parent = workbook_tree.find("main:sheets", MAIN_NAMESPACE)
        if sheets_parent is None:
            return

        for sheet in sheets_parent.findall("main:sheet", MAIN_NAMESPACE):
            name = sheet.attrib.get("name")
            rel_id = sheet.attrib.get(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
            )
            if not name or not rel_id:
                continue
            target = relationship_map.get(rel_id)
            if target:
                self.sheet_targets[name] = target

        styles_tree = ET.fromstring(self.archive.read("xl/styles.xml"))
        fills_parent = styles_tree.find("main:fills", MAIN_NAMESPACE)
        cell_xfs_parent = styles_tree.find("main:cellXfs", MAIN_NAMESPACE)
        if fills_parent is None or cell_xfs_parent is None:
            return

        fills = list(fills_parent.findall("main:fill", MAIN_NAMESPACE))
        fill_colors: Dict[int, Optional[str]] = {}
        for idx, fill in enumerate(fills):
            pattern = fill.find("main:patternFill", MAIN_NAMESPACE)
            if pattern is None:
                fill_colors[idx] = None
                continue
            fg_color = pattern.find("main:fgColor", MAIN_NAMESPACE)
            if fg_color is None:
                fill_colors[idx] = None
            elif "rgb" in fg_color.attrib:
                fill_colors[idx] = fg_color.attrib["rgb"].upper()
            else:
                fill_colors[idx] = None

        self.style_outcome.clear()
        for idx, xf in enumerate(cell_xfs_parent.findall("main:xf", MAIN_NAMESPACE)):
            fill_id = int(xf.attrib.get("fillId", "0"))
            color = fill_colors.get(fill_id)
            if color == "FF92D050":
                outcome = "win"
            elif color == "FFFF0000":
                outcome = "loss"
            else:
                outcome = "pending"
            self.style_outcome[idx] = outcome

    def _get_sheet_style_map(self, sheet_name: str) -> Dict[str, int]:
        assert self.archive is not None

        target = self.sheet_targets.get(sheet_name)
        if not target:
            return {}
        if target in self._style_cache:
            return self._style_cache[target]

        xml_bytes = self.archive.read(f"xl/{target}")
        tree = ET.fromstring(xml_bytes)
        style_map: Dict[str, int] = {}
        for cell in tree.findall(".//main:c", MAIN_NAMESPACE):
            ref = cell.attrib.get("r")
            style = cell.attrib.get("s")
            if ref and style is not None:
                try:
                    style_map[ref] = int(style)
                except ValueError:
                    continue
        self._style_cache[target] = style_map
        return style_map

    def get_cell_outcome(self, sheet_name: str, cell_ref: str) -> str:
        style_map = self._get_sheet_style_map(sheet_name)
        style_idx = style_map.get(cell_ref)
        if style_idx is None:
            return "pending"
        return self.style_outcome.get(style_idx, "pending")


def normalise_schedule_line(value: Any) -> Any:
    if isinstance(value, (int, float)):
        return float(value)
    return value


def _looks_like_schedule_start(date_or_time: Any, team: Any, line: Any) -> bool:
    if isinstance(team, str) and team.strip().lower() in SCHEDULE_HEADER_TOKENS:
        return False

    if isinstance(line, str) and line.strip().lower() in SCHEDULE_HEADER_TOKENS:
        return False

    if isinstance(date_or_time, (datetime, time)):
        return True

    if isinstance(date_or_time, str):
        stripped = date_or_time.strip()
        if stripped:
            lowered = stripped.lower()
            if re.search(r"\b\d{1,2}(:\d{2})?\s*(am|pm)\b", lowered):
                return True
            if re.fullmatch(r"\d{1,2}:\d{2}", stripped):
                return True
            if re.search(r"\b(?:mon|tue|wed|thu|fri|sat|sun)\b", lowered):
                return True
            if re.search(r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\b", lowered):
                return True
            if re.search(r"\d{4}-\d{2}-\d{2}", stripped):
                return True
            if re.search(r"\d{1,2}/\d{1,2}", stripped):
                return True

    if isinstance(team, str) and team.strip() and line not in (None, ""):
        return True

    return False


def _should_ignore_schedule_cells(date_or_time: Any, team: Any, line: Any) -> bool:
    if date_or_time is None and team is None and line is None:
        return True

    if isinstance(team, str) and team.strip().lower() in SCHEDULE_HEADER_TOKENS:
        return True

    if isinstance(line, str) and line.strip().lower() in SCHEDULE_HEADER_TOKENS:
        return True

    if date_or_time is not None and not _looks_like_schedule_start(date_or_time, team, line):
        return True

    return False


def parse_schedule_rows(rows: List[tuple], header_index: int) -> List[Dict[str, Any]]:
    schedule: List[Dict[str, Any]] = []
    column_starts: List[int] = []
    for row in rows[:header_index]:
        row_starts: List[int] = []
        row_length = len(row)
        for col in range(row_length):
            cells = [
                row[col + offset] if col + offset < row_length else None
                for offset in range(SCHEDULE_GROUP_WIDTH)
            ]
            if _looks_like_schedule_start(*cells):
                row_starts.append(col)
        if row_starts:
            column_starts = row_starts
            break

    if not column_starts:
        column_starts = [SCHEDULE_START_COLUMN]

    # ensure deterministic ordering and no duplicates
    column_starts = sorted(dict.fromkeys(column_starts))

    current_rows: Dict[int, Dict[str, Any]] = {start: {} for start in column_starts}

    for row in rows[:header_index]:
        for start in column_starts:
            cells = [
                row[start + offset] if start + offset < len(row) else None
                for offset in range(SCHEDULE_GROUP_WIDTH)
            ]
            if not any(cell is not None for cell in cells):
                continue

            date_or_time, team, line = cells
            entry = current_rows[start]

            if not entry:
                if _should_ignore_schedule_cells(date_or_time, team, line):
                    continue

                current_rows[start] = {
                    "team": serialise_datetime(team) if team is not None else "",
                    "line": normalise_schedule_line(line),
                    "date": serialise_datetime(date_or_time),
                }
            else:
                entry.update(
                    {
                        "opponent": serialise_datetime(team) if team is not None else "",
                        "time": serialise_datetime(date_or_time),
                        "opponent_line": normalise_schedule_line(line),
                    }
                )
                schedule.append(entry.copy())
                current_rows[start] = {}

    for entry in current_rows.values():
        if entry:
            schedule.append(entry.copy())

    filtered: List[Dict[str, Any]] = []
    for item in schedule:
        team = item.get("team")
        opponent = item.get("opponent")
        line = item.get("line")
        opponent_line = item.get("opponent_line")

        has_team = isinstance(team, str) and team.strip() and team.strip().lower() not in SCHEDULE_HEADER_TOKENS
        has_opponent = isinstance(opponent, str) and opponent.strip()
        has_line = line not in (None, "") or opponent_line not in (None, "")

        if has_team or has_opponent or has_line:
            filtered.append(item)

    return filtered


def detect_header_row(rows: List[tuple]) -> Optional[int]:
    for idx, row in enumerate(rows):
        if row and row[0] is None and sum(isinstance(cell, (int, float)) for cell in row) >= 2:
            return idx
    for idx, row in enumerate(rows):
        if any(isinstance(cell, (int, float)) for cell in row):
            return idx
    return None


def parse_week_sheet(
    sheet, style_inspector: Optional[WorkbookStyleInspector] = None
) -> Dict[str, Any]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"name": sheet.title, "players": [], "schedule": [], "confidence_points": []}

    header_index = detect_header_row(rows)
    if header_index is None:
        return {
            "name": sheet.title,
            "players": [],
            "schedule": parse_schedule_rows(rows, 0),
            "confidence_points": [],
        }

    header_row = rows[header_index]

    # Confidence point slots live immediately after the first numeric cell.
    confidence_points: List[int] = []
    start_col: Optional[int] = None
    for idx, value in enumerate(header_row):
        if isinstance(value, (int, float)):
            if start_col is None:
                start_col = idx
            confidence_points.append(int(value))
        elif start_col is not None:
            # stop once we hit a non-numeric cell after the numeric streak
            break

    if start_col is None:
        start_col = 1

    total_col = start_col + len(confidence_points)
    best_bet_time_col = total_col + 3
    best_bet_team_col = best_bet_time_col + 1
    best_bet_line_col = best_bet_team_col + 1

    players: List[Dict[str, Any]] = []
    for row_offset, row in enumerate(rows[header_index + 1 :]):
        player = row[0]
        if not player:
            continue
        selections = []
        computed_total = 0
        has_recorded_result = False
        excel_row = header_index + 2 + row_offset
        for offset, points in enumerate(confidence_points):
            col_idx = start_col + offset
            pick = row[col_idx]
            if pick:
                column_letter = get_column_letter(col_idx + 1)
                cell_ref = f"{column_letter}{excel_row}"
                outcome = (
                    style_inspector.get_cell_outcome(sheet.title, cell_ref)
                    if style_inspector
                    else "pending"
                )
                if outcome in {"win", "loss"}:
                    has_recorded_result = True
                awarded = (
                    int(points)
                    if outcome == "win"
                    else 0
                    if outcome == "loss"
                    else None
                )
                if awarded:
                    computed_total += awarded
                selection: Dict[str, Any] = {
                    "team": pick,
                    "points": int(points),
                    "result": outcome,
                }
                if awarded is not None:
                    selection["awarded_points"] = awarded
                else:
                    selection["awarded_points"] = None
                selections.append(selection)
        total = row[total_col] if total_col < len(row) else None
        if total is None and (has_recorded_result or computed_total):
            total = computed_total
        best_bet = {
            "time": row[best_bet_time_col] if best_bet_time_col < len(row) else None,
            "team": row[best_bet_team_col] if best_bet_team_col < len(row) else None,
            "line": row[best_bet_line_col] if best_bet_line_col < len(row) else None,
        }
        if not any(best_bet.values()):
            best_bet = None
        else:
            best_bet = {
                "time": serialise_datetime(best_bet.get("time")),
                "team": best_bet.get("team"),
                "line": float(best_bet["line"]) if isinstance(best_bet.get("line"), (int, float)) else best_bet.get("line"),
            }
        players.append(
            {
                "name": player,
                "picks": selections,
                "total_points": total,
                "best_bet": best_bet,
            }
        )

    schedule = parse_schedule_rows(rows, header_index)

    return {
        "name": sheet.title,
        "players": players,
        "schedule": schedule,
        "confidence_points": confidence_points,
    }


def parse_standings(sheet) -> Dict[str, Dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}
    header_row = rows[0]
    try:
        player_col = header_row.index("Player")
    except ValueError as exc:
        raise ValueError("Unable to locate Player column in standings") from exc

    week_columns: Dict[str, int] = {}
    for idx, value in enumerate(header_row):
        if isinstance(value, str) and value.startswith("Week"):
            week_columns[value] = idx

    standings: Dict[str, Dict[str, Any]] = {}
    for row in rows[1:]:
        player = row[player_col]
        if not isinstance(player, str):
            continue
        player_data: Dict[str, Any] = {}
        for week, col_idx in week_columns.items():
            score = row[col_idx] if col_idx < len(row) else None
            if score is not None:
                player_data[week] = score
        standings[player] = player_data
    return standings


def main() -> None:
    with WorkbookStyleInspector(GRID_PATH) as style_inspector:
        workbook = openpyxl.load_workbook(GRID_PATH, data_only=True, read_only=True)

        week_sheets = [
            workbook[sheet]
            for sheet in workbook.sheetnames
            if sheet.startswith("Week ")
        ]
        week_sheets.sort(key=lambda ws: int(ws.title.split()[1]))

        weeks_data = [parse_week_sheet(sheet, style_inspector) for sheet in week_sheets]

        standings_sheet = (
            workbook["Standings"] if "Standings" in workbook.sheetnames else None
        )
        standings = parse_standings(standings_sheet) if standings_sheet else {}

        for week in weeks_data:
            week_name = week.get("name")
            if not isinstance(week_name, str):
                continue
            for player in week.get("players", []):
                score = player.get("total_points")
                if not isinstance(score, (int, float)):
                    continue
                player_name = player.get("name")
                if not isinstance(player_name, str):
                    continue
                player_standings = standings.setdefault(player_name, {})
                player_standings[week_name] = score

        dataset = {
            "generatedAt": datetime.now(UTC).isoformat(),
            "weeks": weeks_data,
            "standings": standings,
        }

        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        OUTPUT_PATH.write_text(json.dumps(dataset, indent=2), encoding="utf-8")
        print(f"Wrote data for {len(weeks_data)} weeks to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
